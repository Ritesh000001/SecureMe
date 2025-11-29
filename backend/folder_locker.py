import os
import subprocess
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from config import BASE_DIR

LOCK_LOG = BASE_DIR / "data" / "LockedFolders.xlsx"
LOCK_LOG.parent.mkdir(parents=True, exist_ok=True)


def ensure_log_exists():
    if not LOCK_LOG.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "LockedFolders"
        ws.append(["Folder Path", "Status", "Date-Time"])
        wb.save(LOCK_LOG)


def update_log(folder_path, status):
    ensure_log_exists()
    wb = load_workbook(LOCK_LOG)
    ws = wb.active

    found = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == folder_path:
            row[1].value = status
            row[2].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            found = True
            break

    if not found:
        ws.append([folder_path, status, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    wb.save(LOCK_LOG)


def lock_folder(folder_path):
    if not os.path.exists(folder_path):
        return False, "❌ Folder not found."

    try:
        username = os.getlogin()
        cmd = f'icacls "{folder_path}" /deny {username}:(RX)'
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True)

        if result.returncode == 0:
            update_log(folder_path, "Locked")
            return True, f"✅ Folder locked: {folder_path}"
        else:
            return False, f"⚠️ Failed: {result.stderr.strip()}"

    except Exception as e:
        return False, f"⚠️ Error: {str(e)}"


def unlock_folder(folder_path):
    if not os.path.exists(folder_path):
        return False, "❌ Folder not found."

    try:
        username = os.getlogin()
        cmd = f'icacls "{folder_path}" /remove:d {username}'
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True)

        if result.returncode == 0:
            update_log(folder_path, "Unlocked")
            return True, f"🔓 Folder unlocked: {folder_path}"
        else:
            return False, f"⚠️ Failed: {result.stderr.strip()}"

    except Exception as e:
        return False, f"⚠️ Error: {str(e)}"


def list_locked_folders():
    ensure_log_exists()
    wb = load_workbook(LOCK_LOG)
    ws = wb.active

    folders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        folders.append({
            "path": row[0],
            "status": row[1],
            "datetime": row[2],
        })

    return folders


def count_locked_folders():
    folders = list_locked_folders()
    return sum(1 for f in folders if f["status"] == "Locked")
