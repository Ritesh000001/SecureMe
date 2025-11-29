import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from config import BASE_DIR
from backend.encryption_utils import encrypt_text, decrypt_text

VAULT_FILE = BASE_DIR / "data" / "PasswordVault.xlsx"
VAULT_FILE.parent.mkdir(parents=True, exist_ok=True)

def ensure_vault_exists():
    if not VAULT_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Vault"
        ws.append(["Website", "Name", "Email/Username/Phone", "Password", "Category", "Date"])
        wb.save(VAULT_FILE)

def save_entry(website: str, name: str, contact: str, password: str, category: str) -> tuple[bool, str]:
    ensure_vault_exists()
    try:
        encrypted_pass = encrypt_text(password)
        wb = load_workbook(VAULT_FILE)
        ws = wb.active
        ws.append([website, name, contact, encrypted_pass, category, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        wb.save(VAULT_FILE)

        return True, "✅ Password entry saved successfully."
    except Exception as e:
        return False, f"❌ Error saving entry: {e}"

def list_entries():
    ensure_vault_exists()
    wb = load_workbook(VAULT_FILE)
    ws = wb.active

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        website, name, contact, enc_pass, category, date = row
        try:
            decrypted_pass = decrypt_text(enc_pass)
        except Exception:
            decrypted_pass = "⚠️ Error"
        entries.append({
            "website": website,
            "name": name,
            "contact": contact,
            "password": decrypted_pass,
            "category": category,
            "date": date
        })
    return entries

def count_vault_entries() -> int:
    ensure_vault_exists()
    wb = load_workbook(VAULT_FILE)
    ws = wb.active
    return ws.max_row - 1

def update_entry(entry_id: int, website: str, name: str, contact: str, password: str, category: str) -> tuple[bool, str]:
    """Update an existing vault entry by row index (1-based excluding header)."""
    ensure_vault_exists()
    try:
        wb = load_workbook(VAULT_FILE)
        ws = wb.active

        # +1 to skip header row
        row_idx = entry_id + 1  
        if row_idx > ws.max_row:
            return False, "❌ Entry not found."

        encrypted_pass = encrypt_text(password)
        ws.cell(row=row_idx, column=1).value = website
        ws.cell(row=row_idx, column=2).value = name
        ws.cell(row=row_idx, column=3).value = contact
        ws.cell(row=row_idx, column=4).value = encrypted_pass
        ws.cell(row=row_idx, column=5).value = category
        ws.cell(row=row_idx, column=6).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb.save(VAULT_FILE)
        return True, "✅ Entry updated successfully."
    except Exception as e:
        return False, f"❌ Error updating entry: {e}"

