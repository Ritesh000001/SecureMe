import os
import hashlib
import random
import string
import base64
from datetime import datetime
from cryptography.fernet import Fernet
from openpyxl import Workbook, load_workbook
from pathlib import Path
from config import BASE_DIR

# -----------------------------------------------------------------
# Paths
# -----------------------------------------------------------------
META_FILE = BASE_DIR / "data" / "ONotes.xlsx"
VAULT_KEY_FILE = BASE_DIR / "data" / "vault_master.key"
META_FILE.parent.mkdir(parents=True, exist_ok=True)


# -----------------------------------------------------------------
# Utilities for file-based encryption (used by Notes)
# -----------------------------------------------------------------
def generate_random_key():
    return ''.join(random.choices(string.ascii_letters + string.digits, k=6))


def derive_fernet(user_key: str) -> Fernet:
    digest = hashlib.sha256(user_key.encode("utf-8")).digest()
    b64key = base64.urlsafe_b64encode(digest[:32])
    return Fernet(b64key)


def ensure_meta_exists():
    if not META_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "NotesKeys"
        ws.append(["File Name", "Algorithm", "Hash of Key", "Date-Time", "Key"])
        wb.save(META_FILE)


def upsert_metadata(file_name: str, algo: str, key: str):
    ensure_meta_exists()
    wb = load_workbook(META_FILE)
    ws = wb.active
    key_hash = hashlib.sha256(key.encode("utf-8")).hexdigest()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    found_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == file_name:
            found_row = r
            break

    if found_row:
        ws.cell(row=found_row, column=2).value = algo
        ws.cell(row=found_row, column=3).value = key_hash
        ws.cell(row=found_row, column=4).value = now
        ws.cell(row=found_row, column=5).value = key
    else:
        ws.append([file_name, algo, key_hash, now, key])

    wb.save(META_FILE)


def encrypt_file(file_path: Path, user_key: str):
    fernet = derive_fernet(user_key)
    with open(file_path, "rb") as f:
        data = f.read()
    token = fernet.encrypt(data)
    with open(file_path, "wb") as f:
        f.write(token)
    upsert_metadata(file_path.name, "Fernet", user_key)


def decrypt_file(file_path: Path, user_key: str) -> bool:
    fernet = derive_fernet(user_key)
    with open(file_path, "rb") as f:
        token = f.read()
    try:
        data = fernet.decrypt(token)
    except Exception:
        return False
    with open(file_path, "wb") as f:
        f.write(data)
    return True


# -----------------------------------------------------------------
# Text encryption (used by Password Vault)
# -----------------------------------------------------------------
def _get_vault_key() -> bytes:
    if not VAULT_KEY_FILE.exists():
        key = Fernet.generate_key()
        VAULT_KEY_FILE.write_bytes(key)
        return key
    return VAULT_KEY_FILE.read_bytes()


def encrypt_text(plaintext: str) -> str:
    if not plaintext:
        return ""
    fernet = Fernet(_get_vault_key())
    encrypted = fernet.encrypt(plaintext.encode("utf-8"))
    return encrypted.decode("utf-8")


def decrypt_text(ciphertext: str) -> str:
    if not ciphertext:
        return ""
    fernet = Fernet(_get_vault_key())
    decrypted = fernet.decrypt(ciphertext.encode("utf-8"))
    return decrypted.decode("utf-8")
